import streamlit as st
import pandas as pd
import requests
from openpyxl import load_workbook
import concurrent.futures
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import matplotlib.pyplot as plt
import io
from pypdf import PdfReader
from bs4 import BeautifulSoup
import time
import random

# --- 1. CONFIGURACIÓN DE LA PÁGINA ---
st.set_page_config(page_title="Laboratorio Modular", page_icon="🎛️", layout="wide")

# --- 2. BARRA LATERAL (PANEL DE CONTROL) ---
with st.sidebar:
    st.warning("⚠️ MODO LABORATORIO (PRUEBAS)")
    
    st.header("🎛️ Panel de Control")
    st.markdown("Selecciona qué herramientas activar para optimizar el tiempo de proceso.")
    
    # --- INTERRUPTORES DE HERRAMIENTAS ---
    act_auditoria = st.checkbox("🛠️ Auditar Formatos y Calidad (OCR)", value=True, help="Verifica si es XML, PDF Texto o Imagen.")
    act_busqueda = st.checkbox("🕵️‍♂️ Buscar Contenido (Palabras)", value=True, help="Descarga el archivo y busca texto dentro.")
    
    st.write("---")
    
    # Configuración de Búsqueda (Solo visible si se activa la búsqueda)
    if act_busqueda:
        st.info("ℹ️ Configuración de Búsqueda:")
        texto_busqueda = st.text_area("Palabras a buscar:", value="puente, contrato, licitacion")
        lista_palabras = [p.strip().lower() for p in texto_busqueda.split(',') if p.strip()]
    else:
        lista_palabras = [] # Lista vacía si está desactivado

    st.write("---")
    st.caption("🚀 CONTROL DE VELOCIDAD")
    modo_sigilo = st.checkbox("🐢 Modo Sigilo (Anti-bloqueo)", value=False, help="Reduce la velocidad para evitar bloqueos del servidor.")

    st.write("---")
    st.info("🎓 App desarrollada dentro del trabajo de doctorado del Mtro. Fernando Gamez Reyes.")
    
    if st.button("🔒 Cerrar Sesión"):
        st.session_state.usuario_valido = False
        st.rerun()

# --- 3. SEGURIDAD ---
if "usuario_valido" not in st.session_state:
    st.session_state.usuario_valido = False

if not st.session_state.usuario_valido:
    st.markdown("# 🔒 Acceso Privado - LABORATORIO")
    clave = st.text_input("Contraseña:", type="password")
    if st.button("Entrar"):
        if clave == "Fernando2026":
            st.session_state.usuario_valido = True
            st.rerun()
        else:
            st.error("⛔ Incorrecto")
    st.stop()

# --- 4. LÓGICA DEL SISTEMA ---
def crear_sesion_segura():
    session = requests.Session()
    retry = Retry(total=2, backoff_factor=1, status_forcelist=[500, 502, 503, 504])
    session.mount('http://', HTTPAdapter(max_retries=retry))
    session.mount('https://', HTTPAdapter(max_retries=retry))
    return session

def auditar_archivo(response, url, realizar_busqueda, palabras_clave):
    """
    Realiza la auditoría técnica y/o la búsqueda de contenido según se solicite.
    """
    calidad = "No solicitado"
    hallazgos = "No solicitado"
    texto_extraido = ""
    
    headers = response.headers
    content_type = headers.get('Content-Type', '').lower()
    ext = url.split('.')[-1].lower()
    
    # --- FASE 1: AUDITORÍA TÉCNICA (Si se solicita) ---
    # Detectamos formato y calidad OCR
    es_legible = False
    
    # 1. Datos Estructurados
    formatos_datos = ['xml', 'json', 'rdf', 'csv']
    if any(f in ext for f in formatos_datos) or any(f in content_type for f in formatos_datos):
        calidad = f"✅ Formato Abierto ({ext.upper()})"
        es_legible = True # Podríamos leerlo si quisiéramos
    
    # 2. PDF
    elif 'pdf' in ext or 'application/pdf' in content_type:
        try:
            f = io.BytesIO(response.content)
            reader = PdfReader(f)
            limit = min(3, len(reader.pages)) 
            for i in range(limit):
                page_text = reader.pages[i].extract_text()
                if page_text:
                    texto_extraido += page_text + " "
            
            if len(texto_extraido.strip()) > 5:
                calidad = "✅ PDF Texto (Abierto)"
                es_legible = True
            else:
                calidad = "⚠️ PDF Imagen (Requiere OCR)"
                es_legible = False
        except:
            calidad = "❌ PDF Dañado"
            
    # 3. HTML
    elif 'html' in ext or 'text/html' in content_type:
        try:
            soup = BeautifulSoup(response.content, 'html.parser')
            texto_extraido = soup.get_text()
            calidad = "✅ Sitio Web (HTML)"
            es_legible = True
        except:
            calidad = "⚠️ HTML con errores"
            
    else:
        calidad = f"⚠️ Formato No Estándar ({ext.upper()})"

    # --- FASE 2: BÚSQUEDA DE CONTENIDO (Si se solicita y es legible) ---
    if realizar_busqueda:
        lista_hallazgos = []
        if es_legible and texto_extraido:
            texto_norm = texto_extraido.lower()
            for palabra in palabras_clave:
                if palabra in texto_norm:
                    lista_hallazgos.append(palabra.upper())
            hallazgos = f"✅ ENCONTRADO: {', '.join(lista_hallazgos)}" if lista_hallazgos else "Sin coincidencias"
        elif not es_legible and "PDF Imagen" in calidad:
            hallazgos = "❌ Imposible leer (Es imagen)"
        else:
            hallazgos = "No legible / Sin texto"

    return calidad, hallazgos

def procesar_enlace(datos):
    # Pausa de Sigilo
    if datos['Sigilo']:
        time.sleep(random.uniform(0.5, 2.0))
    
    url = datos['URL Original']
    act_auditoria = datos['Activar Auditoría']
    act_busqueda = datos['Activar Búsqueda']
    palabras = datos['Palabras Clave']
    
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
    session = crear_sesion_segura()
    
    datos['Estado'] = "Desconocido"
    datos['Formato'] = "Off"
    datos['Contenido'] = "Off"
    
    try:
        # OPTIMIZACIÓN INTELIGENTE:
        # Si NO pedimos auditoría NI búsqueda, usamos HEAD (Ultra rápido, no descarga).
        # Si pedimos CUALQUIERA de las dos, necesitamos GET (Descargar archivo).
        necesita_descarga = act_auditoria or act_busqueda
        
        if necesita_descarga:
            response = session.get(url, headers=headers, timeout=15, stream=False)
        else:
            response = session.head(url, headers=headers, timeout=5, allow_redirects=True)
            # Si el servidor rechaza HEAD (405), intentamos GET ligero
            if response.status_code == 405:
                response = session.get(url, headers=headers, timeout=5, stream=True)

        datos['Código'] = response.status_code
        
        if response.status_code == 200:
            datos['Estado'] = "✅ ACTIVO"
            datos['Tipo'] = "Accesible"
            
            # Solo ejecutamos lógica pesada si el usuario activó los interruptores
            calidad = "No analizado"
            hallazgos = "No analizado"
            
            if necesita_descarga:
                # Si se pidió auditoría, se procesa. Si se pidió búsqueda, también.
                # Pasamos 'act_busqueda' para que la función sepa si debe buscar palabras o no.
                res_calidad, res_hallazgos = auditar_archivo(response, url, act_busqueda, palabras)
                
                if act_auditoria:
                    datos['Formato'] = res_calidad
                if act_busqueda:
                    datos['Contenido'] = res_hallazgos
            
        elif response.status_code == 404:
            datos['Estado'] = "❌ ROTO"
            datos['Tipo'] = "Inaccesible"
        else:
            datos['Estado'] = f"⚠️ ({response.status_code})"
            datos['Tipo'] = "Error"
            
    except Exception:
        datos['Estado'] = "💀 ERROR"
        datos['Tipo'] = "Fallo"
        datos['Formato'] = "Error Conexión"
    finally:
        session.close()
    return datos

# --- 5. INTERFAZ PRINCIPAL ---

st.title("🎛️ Laboratorio Modular de Auditoría")

st.markdown("""
**Personaliza tu auditoría según el tiempo disponible:**
* **Solo Verificación:** Ultrarápido. Solo comprueba disponibilidad.
* **+ Auditoría:** Verifica formatos abiertos y calidad OCR.
* **+ Búsqueda:** Análisis profundo de contenido (Mayor tiempo de proceso).
""")

archivo_subido = st.file_uploader("Carga Excel (.xlsx)", type=["xlsx"])

if archivo_subido and st.button("🚀 Iniciar Proceso"):
    wb = load_workbook(archivo_subido, data_only=True)
    lista_trabajo = []
    
    st.write("⚙️ Configurando robots...")
    
    for hoja in wb.sheetnames:
        ws = wb[hoja]
        for row in ws.iter_rows():
            for cell in row:
                url = None
                if cell.hyperlink:
                    url = cell.hyperlink.target
                elif isinstance(cell.value, str) and str(cell.value).startswith(('http', 'https')):
                    url = cell.value
                
                if url:
                    lista_trabajo.append({
                        "Hoja": hoja,
                        "Celda": cell.coordinate,
                        "URL Original": url,
                        "Activar Auditoría": act_auditoria,
                        "Activar Búsqueda": act_busqueda,
                        "Palabras Clave": lista_palabras,
                        "Sigilo": modo_sigilo
                    })
    
    total = len(lista_trabajo)
    if total == 0:
        st.warning("No se encontraron enlaces.")
    else:
        # Lógica de Workers
        if modo_sigilo:
            workers = 2
            mensaje_vel = "🐢 MODO SIGILO ACTIVADO"
        else:
            # Si solo es verificación simple (HEAD), podemos usar muchos más robots porque es muy ligero
            if not act_auditoria and not act_busqueda:
                workers = 12 # ¡Súper Rápido!
                mensaje_vel = "⚡ MODO FLASH (Solo Verificación)"
            else:
                workers = 8
                mensaje_vel = "🚀 MODO ESTÁNDAR (Análisis Completo)"
        
        st.info(f"{mensaje_vel}: Procesando {total} enlaces con {workers} robots...")
        
        barra = st.progress(0)
        estado = st.empty()
        resultados = []
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {executor.submit(procesar_enlace, item): item for item in lista_trabajo}
            completados = 0
            for future in concurrent.futures.as_completed(futures):
                resultados.append(future.result())
                completados += 1
                barra.progress(int((completados/total)*100))
                estado.text(f"Progreso: {completados}/{total}...")
        
        barra.progress(100)
        estado.success("✅ Proceso Finalizado")
        df = pd.DataFrame(resultados)
        
        # --- PESTAÑAS DINÁMICAS ---
        # Mostramos pestañas según lo que se activó
        tabs_titulos = ["📄 Resultados Generales", "📊 Gráficos"]
        if act_auditoria:
            tabs_titulos.insert(1, "🛠️ Detalles Técnicos")
        if act_busqueda:
            tabs_titulos.insert(2, "🕵️‍♂️ Hallazgos de Contenido")
            
        tabs = st.tabs(tabs_titulos)
        
        # 1. General
        with tabs[0]:
            st.dataframe(df)
            st.download_button("Descargar CSV", df.to_csv(index=False).encode('utf-8'), "auditoria_modular.csv")
            
        # Pestañas condicionales
        idx = 1
        if act_auditoria:
            with tabs[idx]:
                st.subheader("Análisis de Formatos")
                c1, c2 = st.columns(2)
                c1.warning("⚠️ Requieren OCR (Imagen):")
                c1.dataframe(df[df['Formato'].
